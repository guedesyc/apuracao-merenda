import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DB_FILE = ROOT / "data" / "db.json"
TEMPLATE = ROOT / "data" / "templates" / "Pasta1.xlsx"
EXPORT_DIR = ROOT / "data" / "exports"


def has_filled_quantities(entry):
    quantities = entry.get("quantities") or {}
    if not quantities:
        return False
    for value in quantities.values():
        if value in ("", None):
            return False
        try:
            float(str(value).replace(",", "."))
        except ValueError:
            return False
    return True


def quantity_to_int(value):
    return int(float(str(value).replace(",", ".")))


def main():
    month = sys.argv[1] if len(sys.argv) > 1 else "2026-07"
    if len(month) != 7:
        raise SystemExit("Competencia invalida. Use AAAA-MM.")

    db = json.loads(DB_FILE.read_text(encoding="utf-8"))
    cards = {card["id"]: card for card in db["cards"]}
    totals = defaultdict(lambda: defaultdict(int))
    nutritionists = defaultdict(set)

    for entry in db["entries"]:
        if not entry.get("date", "").startswith(month):
            continue
        if entry.get("status") == "not_served":
            if not entry.get("reason"):
                continue
            continue
        if not has_filled_quantities(entry):
            continue
        school_id = entry["schoolId"]
        nutritionists[school_id].add(entry.get("nutritionistName", ""))
        for card_id, qty in entry.get("quantities", {}).items():
            totals[school_id][card_id] += quantity_to_int(qty or 0)

    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.active
    school_by_row = {school["row"]: school for school in db["schools"]}

    for row in range(5, 170):
        school = school_by_row.get(row)
        if not school:
            continue
        ws.cell(row, 1).value = ", ".join(sorted(filter(None, nutritionists[school["id"]]))) or None
        for card in db["cards"]:
            ws.cell(row, card["column"]).value = totals[school["id"]].get(card["id"], 0)

    for card in db["cards"]:
        col_letter = ws.cell(1, card["column"]).column_letter
        ws.cell(173, card["column"]).value = f"={col_letter}171*{col_letter}172"

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"apuracao-consolidada-{month}-{stamp}.xlsx"
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = EXPORT_DIR / filename
    wb.save(output_path)

    db.setdefault("exports", []).append(
        {
            "month": month,
            "filename": filename,
            "createdAt": datetime.now().isoformat(timespec="seconds"),
            "rows": len(school_by_row),
        }
    )
    DB_FILE.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "filename": filename, "url": f"/exports/{filename}"}))


if __name__ == "__main__":
    main()
