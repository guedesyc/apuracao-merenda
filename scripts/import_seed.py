import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "data" / "templates" / "Pasta1.xlsx"
ADDRESSES = ROOT / "data" / "templates" / "ENDERECOS.xlsx"
DB_FILE = ROOT / "data" / "db.json"

ROUTE_NAMES = {
    "CAJAZEIRAS",
    "SÃO CAETANO",
    "SAO CAETANO",
    "LIBERDADE",
    "CENTRO",
    "CIDADE BAIXA",
    "SUBURBIO",
    "SUBÚRBIO",
    "SUBURBIO I",
    "SUBURBIO II",
    "COUTOS",
}


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def normalize(value):
    text = clean(value).upper()
    text = re.sub(r"^\d+\s*-\s*", "", text)
    replacements = str.maketrans("ÁÀÃÂÉÊÍÓÔÕÚÜÇªº", "AAAAEEIOOOUUCAO")
    text = text.translate(replacements)
    return re.sub(r"[^A-Z0-9]+", " ", text).strip()


def slug(value):
    text = normalize(value).lower()
    return re.sub(r"\s+", ".", text).strip(".")


def parse_addresses():
    if not ADDRESSES.exists():
        return {}
    wb = openpyxl.load_workbook(ADDRESSES, read_only=True, data_only=True)
    ws = wb["ENDEREÇOS"] if "ENDEREÇOS" in wb.sheetnames else wb[wb.sheetnames[0]]
    current_route = ""
    by_school = {}

    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        gre, company, school, address = [clean(item) for item in row]
        if not any([gre, company, school, address]):
            continue
        possible_route = normalize(gre or school)
        if possible_route in {normalize(item) for item in ROUTE_NAMES} and not address:
            current_route = clean(gre or school).upper()
            continue
        route = gre or current_route
        if not school or normalize(school) in {normalize(item) for item in ROUTE_NAMES}:
            if gre:
                current_route = gre.upper()
            continue
        by_school[normalize(school)] = {
            "route": route.upper() if route else "SEM ROTA",
            "company": company,
            "address": address,
        }
        if gre:
            current_route = gre.upper()
    return by_school


def main():
    wb = openpyxl.load_workbook(TEMPLATE, read_only=True, data_only=False)
    ws = wb.active
    address_map = parse_addresses()

    cards = []
    for idx, col in enumerate(range(3, 14), start=1):
        cards.append(
            {
                "id": f"card{idx}",
                "number": idx,
                "label": clean(ws.cell(3, col).value) or f"CARD {idx}",
                "description": clean(ws.cell(4, col).value),
                "price": float(ws.cell(172, col).value or 0),
                "column": col,
            }
        )

    nutritionist_names = []
    for row in range(5, 170):
        name = clean(ws.cell(row, 1).value)
        if name and name not in nutritionist_names:
            nutritionist_names.append(name)

    nutritionist_users = [
        {
            "id": f"user-{slug(name)}",
            "name": name,
            "username": slug(name),
            "password": "123",
            "role": "nutritionist",
            "active": True,
        }
        for name in nutritionist_names
    ]
    user_id_by_name = {user["name"]: user["id"] for user in nutritionist_users}

    schools = []
    for row in range(5, 170):
        nutritionist_name = clean(ws.cell(row, 1).value)
        raw_name = clean(ws.cell(row, 2).value)
        if not raw_name:
            continue
        match = address_map.get(normalize(raw_name), {})
        assigned_user_id = user_id_by_name.get(nutritionist_name)
        schools.append(
            {
                "id": f"school-{row}",
                "row": row,
                "name": raw_name,
                "shortName": re.sub(r"^\d+\s*-\s*", "", raw_name).strip(),
                "route": match.get("route", "SEM ROTA"),
                "company": match.get("company", ""),
                "address": match.get("address", ""),
                "active": True,
                "nutritionistIds": [assigned_user_id] if assigned_user_id else [],
            }
        )

    db = {
        "version": 1,
        "createdAt": "",
        "updatedAt": "",
        "settings": {
            "currentMonth": "2026-07",
            "reasons": ["Sem aula", "Segurança", "Greve", "Feriado", "Outro"],
            "workingDaysByMonth": {"2026-07": 22},
        },
        "cards": cards,
        "schools": schools,
        "users": [
            {"id": "admin", "name": "Coordenação", "username": "admin", "password": "admin", "role": "admin", "active": True},
            *nutritionist_users,
        ],
        "entries": [],
        "closures": [],
        "exports": [],
    }
    DB_FILE.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Base criada em {DB_FILE} com {len(schools)} escolas e {len(cards)} cards.")


if __name__ == "__main__":
    main()
